"""
UPANDRUNNING — Patient Services Daily Dashboard
Single-file Flask application (Python 3).

A standalone Python version of the dashboard. Every change (counters, patient
cancellations, notes, phone numbers, activity log) is saved immediately to a
local SQLite database (autosave). Supports importing the daily cancellations
report from CSV (built-in) and XLSX (if openpyxl is installed).

Requirements:
    pip install flask
    (optional, for .xlsx import) pip install openpyxl

Run:
    python app.py
Then open http://127.0.0.1:5000 in your browser.

NOTE: This is a self-contained app with its own SQLite storage. It is NOT
connected to the Base44 backend / cloud database — data lives in the local
"dashboard.db" file next to this script.
"""

import os
import json
import sqlite3
import csv
import io
import re
from datetime import datetime, date
from flask import Flask, request, jsonify, Response

app = Flask(__name__)
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dashboard.db")

LOCATIONS = [
    {"key": "alwasl", "name": "Al Wasl Road"},
    {"key": "difc", "name": "DIFC"},
    {"key": "egc", "name": "Emirates Golf Club"},
    {"key": "jge", "name": "Jumeirah Golf Estates"},
    {"key": "szr", "name": "SZR Studio Republik"},
]

COUNTER_DEFS = [
    {"field": "phone", "name": "New booking \u2014 Phone"},
    {"field": "whatsapp", "name": "New booking \u2014 WhatsApp"},
    {"field": "referral", "name": "Referral booked"},
    {"field": "waitlist", "name": "Waiting list slot filled"},
]

CENTER_MAP = {
    "sports": "alwasl",
    "al wasl": "alwasl",
    "al wasl road": "alwasl",
    "difc": "difc",
    "egc": "egc",
    "emirates golf club": "egc",
    "jge": "jge",
    "jumeirah golf estates": "jge",
    "szr": "szr",
    "szr studio republik": "szr",
    "studio republik": "szr",
}


# --------------------------------------------------------------------------- #
# Database
# --------------------------------------------------------------------------- #
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """CREATE TABLE IF NOT EXISTS cancellation (
            id TEXT PRIMARY KEY,
            date TEXT NOT NULL,
            location TEXT NOT NULL,
            name TEXT NOT NULL,
            mr_no TEXT DEFAULT '',
            consultant TEXT DEFAULT '',
            reason TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            note TEXT DEFAULT '',
            recouped INTEGER DEFAULT 0,
            created_at TEXT DEFAULT ''
        )"""
    )
    cur.execute(
        """CREATE TABLE IF NOT EXISTS daily_counter (
            id TEXT PRIMARY KEY,
            date TEXT NOT NULL,
            location TEXT NOT NULL,
            phone INTEGER DEFAULT 0,
            whatsapp INTEGER DEFAULT 0,
            referral INTEGER DEFAULT 0,
            waitlist INTEGER DEFAULT 0,
            UNIQUE(date, location)
        )"""
    )
    cur.execute(
        """CREATE TABLE IF NOT EXISTS activity_log (
            id TEXT PRIMARY KEY,
            date TEXT NOT NULL,
            location TEXT NOT NULL,
            label TEXT NOT NULL,
            ts TEXT NOT NULL
        )"""
    )
    conn.commit()
    conn.close()


def new_id():
    import uuid
    return uuid.uuid4().hex


def now_iso():
    return datetime.now().isoformat(timespec="seconds")


def ensure_counters(date_str):
    conn = get_db()
    cur = conn.cursor()
    for loc in LOCATIONS:
        cur.execute(
            "INSERT OR IGNORE INTO daily_counter (id, date, location, phone, whatsapp, referral, waitlist) VALUES (?,?,?,?,0,0,0)",
            (new_id(), date_str, loc["key"]),
        )
    conn.commit()
    conn.close()


# --------------------------------------------------------------------------- #
# Data loaders
# --------------------------------------------------------------------------- #
def load_cancellations(date_str):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM cancellation WHERE date=? ORDER BY created_at DESC", (date_str,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def load_counters(date_str):
    ensure_counters(date_str)
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM daily_counter WHERE date=?", (date_str,)
    ).fetchall()
    conn.close()
    return {r["location"]: dict(r) for r in rows}


def load_logs(date_str):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM activity_log WHERE date=? ORDER BY ts DESC LIMIT 500", (date_str,)
    ).fetchall()
    conn.close()
    out = {}
    for r in rows:
        out.setdefault(r["location"], []).append(dict(r))
    return out


def add_log(conn, date_str, loc_key, label):
    conn.execute(
        "INSERT INTO activity_log (id, date, location, label, ts) VALUES (?,?,?,?,?)",
        (new_id(), date_str, loc_key, label, now_iso()),
    )


def clean_name(n):
    n = re.sub(r"^(mrs|miss|mr|ms|dr)\.?\s*", "", n or "", flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", n).strip()


def normalize_header(h):
    return re.sub(r"[^a-z]", "", (h or "").lower())


def parse_csv_text(text):
    reader = csv.reader(io.StringIO(text))
    return [row for row in reader if any(cell.strip() for cell in row)]


def parse_xlsx_bytes(data):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("XLSX import requires the 'openpyxl' package (pip install openpyxl).")
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    return [[("" if c is None else str(c)) for c in row] for row in ws.iter_rows(values_only=True)]


# --------------------------------------------------------------------------- #
# Routes — API (autosave)
# --------------------------------------------------------------------------- #
@app.route("/api/data")
def api_data():
    date_str = request.args.get("date")
    if not date_str:
        return jsonify({"error": "date required"}), 400
    return jsonify(
        {
            "cancellations": load_cancellations(date_str),
            "counters": load_counters(date_str),
            "logs": load_logs(date_str),
        }
    )


@app.route("/api/counter", methods=["POST"])
def api_counter():
    body = request.get_json()
    date_str, loc_key, field, delta = body["date"], body["location"], body["field"], body["delta"]
    conn = get_db()
    ensure_counters(date_str)
    row = conn.execute(
        "SELECT * FROM daily_counter WHERE date=? AND location=?", (date_str, loc_key)
    ).fetchone()
    new_val = max(0, (row[field] or 0) + delta)
    conn.execute(
        "UPDATE daily_counter SET {}=? WHERE id=?".format(field), (new_val, row["id"])
    )
    if delta > 0:
        label = next((c["name"] for c in COUNTER_DEFS if c["field"] == field), field) + " +1"
        add_log(conn, date_str, loc_key, label)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "value": new_val})


@app.route("/api/patients", methods=["POST"])
def api_add_patients():
    body = request.get_json()
    date_str, loc_key, text = body["date"], body["location"], body["text"]
    names = [clean_name(n) for n in text.split("\n") if n.strip()]
    names = [n for n in names if n]
    conn = get_db()
    created = []
    ts = now_iso()
    for name in names:
        pid = new_id()
        conn.execute(
            "INSERT INTO cancellation (id, date, location, name, mr_no, consultant, reason, phone, note, recouped, created_at) VALUES (?,?,?,?,?,?,?,?,?,0,?)",
            (pid, date_str, loc_key, name, "", "", "", "", "", ts),
        )
        created.append(pid)
    if created:
        add_log(conn, date_str, loc_key, "Added {} patient{} from cancellations report".format(len(created), "s" if len(created) != 1 else ""))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "added": len(created)})


@app.route("/api/patient/<pid>/toggle", methods=["POST"])
def api_toggle(pid):
    body = request.get_json()
    date_str = body["date"]
    conn = get_db()
    row = conn.execute("SELECT * FROM cancellation WHERE id=?", (pid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "not found"}), 404
    new_val = 0 if row["recouped"] else 1
    conn.execute("UPDATE cancellation SET recouped=? WHERE id=?", (new_val, pid))
    add_log(conn, date_str, row["location"], ("Recouped \u2014 " if new_val else "Unmarked \u2014 ") + row["name"])
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "recouped": bool(new_val)})


@app.route("/api/patient/<pid>/note", methods=["POST"])
def api_note(pid):
    note = request.get_json().get("note", "")
    conn = get_db()
    conn.execute("UPDATE cancellation SET note=? WHERE id=?", (note, pid))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/patient/<pid>/phone", methods=["POST"])
def api_phone(pid):
    phone = request.get_json().get("phone", "")
    conn = get_db()
    conn.execute("UPDATE cancellation SET phone=? WHERE id=?", (phone, pid))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/patient/<pid>", methods=["DELETE"])
def api_remove(pid):
    body = request.get_json() or {}
    date_str = body.get("date", "")
    conn = get_db()
    row = conn.execute("SELECT * FROM cancellation WHERE id=?", (pid,)).fetchone()
    conn.execute("DELETE FROM cancellation WHERE id=?", (pid,))
    if row:
        add_log(conn, date_str, row["location"], "Removed \u2014 " + row["name"])
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


def import_rows(date_str, rows):
    header_idx = -1
    header_fields = None
    for i, r in enumerate(rows):
        norm = [normalize_header(c) for c in r]
        if any(h == "patientname" for h in norm):
            header_idx = i
            header_fields = r
            break
    if header_idx == -1:
        return {"error": "Could not find a \u201CPatient Name\u201D column. Make sure the header row is intact."}

    norm = [normalize_header(h) for h in header_fields]
    phone_idx = next((idx for idx, h in enumerate(norm) if "mobile" in h or "phone" in h or "contact" in h or "tel" in h), -1)
    col = {
        "center": norm.index("centername") if "centername" in norm else -1,
        "mrno": norm.index("mrno") if "mrno" in norm else -1,
        "consultant": norm.index("consultant") if "consultant" in norm else -1,
        "patient": norm.index("patientname"),
        "reason": norm.index("cancelreason") if "cancelreason" in norm else -1,
        "phone": phone_idx,
        "last": len(header_fields) - 1,
    }

    conn = get_db()
    existing = [dict(r) for r in conn.execute("SELECT * FROM cancellation WHERE date=?", (date_str,)).fetchall()]
    per_new, per_dup, unmapped = {}, {}, {}
    total_new = total_dup = total_unmapped = 0
    to_create = []
    ts = now_iso()

    for r in rows[header_idx + 1:]:
        fields = [(str(c) if c is not None else "").strip() for c in r]
        if len(fields) < 2 or not any(fields):
            continue
        center_raw = fields[col["center"]] if col["center"] >= 0 else ""
        patient_raw = fields[col["patient"]]
        if not patient_raw:
            continue
        loc_key = CENTER_MAP.get(center_raw.lower())
        if not loc_key:
            if center_raw:
                unmapped[center_raw] = unmapped.get(center_raw, 0) + 1
                total_unmapped += 1
            continue
        name = clean_name(patient_raw)
        if not name:
            continue
        mr_no = fields[col["mrno"]] if col["mrno"] >= 0 else ""
        consultant = fields[col["consultant"]] if col["consultant"] >= 0 else ""
        reason = fields[col["reason"]] if col["reason"] >= 0 else ""
        phone = fields[col["phone"]] if col["phone"] >= 0 else ""
        booking_status = fields[col["last"]].lower()
        recouped = 1 if booking_status == "booked" else 0

        is_dup = any(
            (mr_no and p["mr_no"] and p["mr_no"] == mr_no) or (p["name"].lower() == name.lower())
            for p in existing
        )
        if is_dup:
            per_dup[loc_key] = per_dup.get(loc_key, 0) + 1
            total_dup += 1
            continue

        to_create.append((new_id(), date_str, loc_key, name, mr_no, consultant, reason, phone, "", recouped, ts))
        per_new[loc_key] = per_new.get(loc_key, 0) + 1
        total_new += 1

    if to_create:
        conn.executemany(
            "INSERT INTO cancellation (id, date, location, name, mr_no, consultant, reason, phone, note, recouped, created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            to_create,
        )
    for loc_key, n in per_new.items():
        add_log(conn, date_str, loc_key, "Imported {} patient{} from cancellations report".format(n, "s" if n != 1 else ""))
    conn.commit()
    conn.close()
    return {"totalNew": total_new, "totalDup": total_dup, "totalUnmapped": total_unmapped, "perLocationNew": per_new, "unmapped": unmapped}


@app.route("/api/import", methods=["POST"])
def api_import():
    date_str = request.form.get("date")
    if not date_str:
        return jsonify({"error": "date required"}), 400

    if "file" in request.files:
        f = request.files["file"]
        lower = f.filename.lower()
        raw = f.read()
        try:
            if lower.endswith((".xlsx", ".xls")):
                rows = parse_xlsx_bytes(raw)
            else:
                rows = parse_csv_text(raw.decode("utf-8-sig"))
        except Exception as e:
            return jsonify({"error": "Could not read that file ({}). Try re-exporting, or paste it in instead.".format(e)}), 422
        return jsonify(import_rows(date_str, rows))

    text = request.form.get("text", "")
    if text.strip():
        return jsonify(import_rows(date_str, parse_csv_text(text)))

    return jsonify({"error": "No file or text provided."}), 400


# --------------------------------------------------------------------------- #
# Routes — UI
# --------------------------------------------------------------------------- #
INDEX_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>UPANDRUNNING Patient Services Dashboard</title>
<style>
  :root { --teal:#006272; --teal-dark:#004a56; --light-teal:#E6EEEF; --orange:#E87722; --text:#333; --muted:#777; --border:#ddd; --bg:#FAFAFA; }
  * { box-sizing:border-box; }
  body { margin:0; font-family:-apple-system,"Segoe UI",Arial,sans-serif; background:var(--bg); color:var(--text); }
  #app { max-width:1100px; margin:0 auto; padding:20px 20px 60px; }
  .header { display:flex; justify-content:space-between; align-items:flex-start; flex-wrap:wrap; gap:12px; margin-bottom:18px; }
  .brand .name { font-weight:800; letter-spacing:.5px; color:var(--teal); font-size:20px; }
  .brand .subtitle { color:var(--muted); font-size:13px; letter-spacing:.3px; text-transform:uppercase; margin-top:2px; }
  .date-picker { display:flex; align-items:center; gap:8px; }
  .date-picker button { border:1px solid var(--border); background:#fff; border-radius:6px; width:32px; height:32px; font-size:16px; cursor:pointer; color:var(--teal); }
  .date-picker button:hover { background:var(--light-teal); }
  .date-picker input[type=date] { border:1px solid var(--border); border-radius:6px; padding:6px 10px; font-size:14px; }
  .today-btn { border:1px solid var(--teal); color:var(--teal); background:#fff; border-radius:6px; padding:6px 10px; font-size:13px; cursor:pointer; }
  .today-btn:hover { background:var(--light-teal); }
  .tabs { display:flex; flex-wrap:wrap; gap:6px; margin-bottom:20px; border-bottom:2px solid var(--border); padding-bottom:10px; }
  .tab { padding:8px 14px; border-radius:20px; font-size:13px; font-weight:600; cursor:pointer; border:1px solid var(--border); background:#fff; color:var(--muted); white-space:nowrap; }
  .tab.active { background:var(--teal); border-color:var(--teal); color:#fff; }
  .tab:hover:not(.active) { background:var(--light-teal); }
  .cards-row { display:grid; grid-template-columns:repeat(auto-fit,minmax(160px,1fr)); gap:12px; margin-bottom:20px; }
  .card { background:#fff; border:1px solid var(--border); border-radius:10px; padding:14px 16px; }
  .card .label { font-size:12px; color:var(--muted); text-transform:uppercase; letter-spacing:.4px; margin-bottom:6px; }
  .card .value { font-size:28px; font-weight:700; color:var(--teal); }
  .card .value.orange { color:var(--orange); }
  .card .sub { font-size:12px; color:var(--muted); margin-top:4px; }
  .section-title { font-size:14px; font-weight:700; color:var(--teal); text-transform:uppercase; letter-spacing:.4px; margin:24px 0 10px; }
  .counter-grid { display:grid; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); gap:12px; }
  .counter-card { background:#fff; border:1px solid var(--border); border-radius:10px; padding:14px 16px; display:flex; flex-direction:column; gap:10px; }
  .counter-top { display:flex; justify-content:space-between; align-items:center; }
  .counter-top .count { font-size:24px; font-weight:700; color:var(--teal); }
  .counter-buttons { display:flex; gap:8px; }
  .counter-buttons button { flex:1; border:none; border-radius:8px; padding:10px 0; font-size:16px; font-weight:700; cursor:pointer; }
  .btn-plus { background:var(--teal); color:#fff; } .btn-plus:hover { background:var(--teal-dark); }
  .btn-minus { background:var(--light-teal); color:var(--teal); } .btn-minus:hover { background:#d3e2e4; }
  .patient-panel { background:#fff; border:1px solid var(--border); border-radius:10px; padding:14px 16px; margin-bottom:20px; }
  .patient-panel-head .title { font-size:13px; font-weight:700; color:var(--teal); text-transform:uppercase; letter-spacing:.4px; }
  .paste-row { display:flex; gap:8px; margin-bottom:12px; }
  .paste-row textarea { flex:1; border:1px solid var(--border); border-radius:6px; padding:8px 10px; font-size:13px; resize:vertical; min-height:40px; }
  .paste-row button, .add-one-btn { border:none; background:var(--teal); color:#fff; border-radius:6px; padding:0 16px; font-size:13px; font-weight:600; cursor:pointer; }
  .paste-row button:hover, .add-one-btn:hover { background:var(--teal-dark); }
  .paste-hint { font-size:11px; color:var(--muted); margin:-6px 0 12px; }
  .patient-list { display:flex; flex-direction:column; gap:6px; }
  .patient-row { display:flex; align-items:flex-start; gap:10px; padding:8px 10px; border:1px solid var(--border); border-radius:8px; background:var(--bg); }
  .patient-row.recouped { background:var(--light-teal); border-color:#bcdadd; }
  .patient-row input[type=checkbox] { width:18px; height:18px; cursor:pointer; flex-shrink:0; margin-top:2px; }
  .patient-row .p-name { flex:1; font-size:13px; font-weight:600; }
  .patient-row.recouped .p-name { text-decoration:line-through; color:var(--muted); }
  .patient-row .p-note, .patient-row .p-phone { font-size:12px; border:1px solid transparent; background:transparent; padding:3px 6px; border-radius:6px; }
  .patient-row .p-note { color:var(--muted); width:160px; flex-shrink:0; }
  .patient-row .p-phone { color:var(--text); width:118px; flex-shrink:0; }
  .patient-row .p-phone:hover, .patient-row .p-phone:focus { border-color:var(--border); background:#fff; }
  .patient-row .p-call { flex-shrink:0; display:inline-flex; align-items:center; justify-content:center; width:26px; height:26px; border-radius:50%; background:var(--teal); color:#fff; text-decoration:none; font-size:13px; }
  .patient-row .p-call:hover { background:var(--teal-dark); }
  .patient-row .p-remove { border:none; background:transparent; color:var(--muted); cursor:pointer; font-size:15px; }
  .patient-row .p-remove:hover { color:var(--orange); }
  .log-box { background:#fff; border:1px solid var(--border); border-radius:10px; padding:4px 16px; max-height:260px; overflow-y:auto; }
  .log-row { display:flex; gap:12px; padding:8px 0; border-bottom:1px solid #f0f0f0; font-size:13px; }
  .log-row:last-child { border-bottom:none; }
  .log-time { color:var(--muted); min-width:64px; }
  table.summary { width:100%; border-collapse:collapse; background:#fff; border:1px solid var(--border); border-radius:10px; overflow:hidden; }
  table.summary th, table.summary td { padding:10px 12px; text-align:center; font-size:13px; border-bottom:1px solid #f0f0f0; }
  table.summary th { background:var(--teal); color:#fff; font-weight:600; text-transform:uppercase; font-size:11px; letter-spacing:.3px; }
  table.summary td:first-child, table.summary th:first-child { text-align:left; }
  table.summary tr.total-row td { font-weight:700; background:var(--light-teal); }
  .loading { text-align:center; color:var(--muted); padding:40px 0; font-size:14px; }
  .footer-note { margin-top:28px; font-size:12px; color:var(--muted); text-align:center; }
</style>
</head>
<body>
<div id="app"><div class="loading">Loading dashboard...</div></div>
<script>
const LOCATIONS = __LOCATIONS__;
const COUNTER_DEFS = __COUNTER_DEFS__;
let state = { date: todayStr(), activeTab: "all", data: null, importSummary: null };

function todayStr(){ const d=new Date(); return d.getFullYear()+"-"+String(d.getMonth()+1).padStart(2,"0")+"-"+String(d.getDate()).padStart(2,"0"); }
function fmtTime(ts){ const d=new Date(ts); let h=d.getHours(); const m=String(d.getMinutes()).padStart(2,"0"); const ap=h>=12?"PM":"AM"; h=h%12; if(!h)h=12; return h+":"+m+" "+ap; }
function fmtDateHuman(s){ const d=new Date(s+"T00:00:00"); return d.toLocaleDateString("en-GB",{weekday:"long",day:"numeric",month:"long",year:"numeric"}); }
function esc(s){ const d=document.createElement("div"); d.textContent=s==null?"":String(s); return d.innerHTML; }
function newTotal(c){ return (c.phone||0)+(c.whatsapp||0)+(c.referral||0)+(c.waitlist||0); }

async function api(path, opts){ const res=await fetch(path, opts); if(!res.ok){ const e=await res.json().catch(()=>({})); throw new Error(e.error||("Request failed ("+res.status+")")); } return res.json(); }
async function load(){ state.data = await api("/api/data?date="+state.date); render(); }

function shiftDate(days){ const d=new Date(state.date+"T00:00:00"); d.setDate(d.getDate()+days); state.date=d.toISOString().slice(0,10); refresh(); }
function onDateChange(v){ if(!v)return; state.date=v; refresh(); }
function goToday(){ state.date=todayStr(); refresh(); }
function refresh(){ document.getElementById("app").innerHTML='<div class="loading">Loading dashboard...</div>'; load(); }
function setActiveTab(k){ state.activeTab=k; render(); }

async function adjustCounter(locKey, field, delta){
  try{ await api("/api/counter",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({date:state.date,location:locKey,field,delta})}); await load(); }catch(e){ alert(e.message); }
}
async function addPatients(locKey, text){
  try{ await api("/api/patients",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({date:state.date,location:locKey,text})}); await load(); }catch(e){ alert(e.message); }
}
async function submitPaste(locKey){ const a=document.getElementById("paste-"+locKey); if(!a||!a.value.trim())return; const t=a.value; a.value=""; await addPatients(locKey,t); }
async function addSingle(locKey){ const i=document.getElementById("single-"+locKey); if(!i||!i.value.trim())return; const t=i.value; i.value=""; await addPatients(locKey,t); }
async function toggle(pid){ try{ await api("/api/patient/"+pid+"/toggle",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({date:state.date})}); await load(); }catch(e){ alert(e.message); } }
async function saveNote(pid,val){ await api("/api/patient/"+pid+"/note",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({note:val})}); }
async function savePhone(pid,val){ await api("/api/patient/"+pid+"/phone",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({phone:val})}); await load(); }
async function removePatient(pid){ if(!confirm("Remove this patient?"))return; try{ await api("/api/patient/"+pid,{method:"DELETE",headers:{"Content-Type":"application/json"},body:JSON.stringify({date:state.date})}); await load(); }catch(e){ alert(e.message); } }

async function importFile(file){
  if(!file)return;
  const fd=new FormData(); fd.append("file",file); fd.append("date",state.date);
  try{ state.importSummary=await api("/api/import",{method:"POST",body:fd}); await load(); render(); }catch(e){ state.importSummary={error:e.message}; render(); }
}
async function importPaste(){
  const a=document.getElementById("csv-area"); if(!a||!a.value.trim())return;
  const fd=new FormData(); fd.append("text",a.value); fd.append("date",state.date); a.value="";
  try{ state.importSummary=await api("/api/import",{method:"POST",body:fd}); await load(); render(); }catch(e){ state.importSummary={error:e.message}; render(); }
}
function dismissSummary(){ state.importSummary=null; render(); }

function recoupCount(cancs){ return cancs.filter(p=>p.recouped).length; }
function recoupRate(cancs){ const t=cancs.length; return t?Math.round(recoupCount(cancs)/t*100):null; }

function render(){
  const app=document.getElementById("app"); const all=state.activeTab==="all"; let h="";
  h+='<div class="header"><div class="brand"><div class="name">UPANDRUNNING</div><div class="subtitle">Patient Services Daily Dashboard</div></div>';
  h+='<div class="date-picker"><button onclick="shiftDate(-1)">&#8249;</button><input type="date" value="'+state.date+'" onchange="onDateChange(this.value)"/><button onclick="shiftDate(1)">&#8250;</button><button class="today-btn" onclick="goToday()">Today</button></div></div>';
  h+='<div style="font-size:13px;color:var(--muted);margin-bottom:14px;">'+fmtDateHuman(state.date)+'</div>';
  h+='<div class="tabs"><div class="tab '+(all?"active":"")+'" onclick="setActiveTab(\'all\')">All Locations</div>';
  LOCATIONS.forEach(l=>{ h+='<div class="tab '+(state.activeTab===l.key?"active":"")+'" onclick="setActiveTab(\''+l.key+'\')">'+l.name+'</div>'; });
  h+='</div>';
  h+= all? renderAll() : renderLocation(state.activeTab);
  h+='<div class="footer-note">Single-file Python app — changes save automatically to a local SQLite database.</div>';
  app.innerHTML=h;
}

function card(label,value,sub,orange){ return '<div class="card"><div class="label">'+label+'</div><div class="value'+(orange?' orange':'')+'">'+value+'</div>'+(sub?'<div class="sub">'+sub+'</div>':'')+'</div>'; }

function renderAll(){
  let h="";
  if(state.importSummary){ const s=state.importSummary; h+='<div style="background:var(--light-teal);border:1px solid #bcdadd;border-radius:10px;padding:12px 16px;margin-bottom:16px;font-size:13px;">';
    if(s.error){ h+='<div style="color:var(--orange);font-weight:600;">Import did not run</div><div>'+esc(s.error)+'</div>'; }
    else{ h+='<div style="font-weight:700;color:var(--teal);margin-bottom:6px;">Import complete</div><div>'+s.totalNew+' patient'+(s.totalNew===1?'':'s')+' added'+(s.totalDup?(", "+s.totalDup+" already on today's list (skipped)"):'')+'.</div>';
      if(s.unmapped&&Object.keys(s.unmapped).length){ h+='<div style="margin-top:6px;color:var(--orange);">Not imported: '+Object.keys(s.unmapped).map(k=>k+" ("+s.unmapped[k]+")").join(", ")+'</div>'; } }
    h+='<button onclick="dismissSummary()" style="margin-top:8px;border:none;background:transparent;color:var(--teal);font-size:12px;font-weight:600;cursor:pointer;">Dismiss</button></div>'; }

  h+='<div class="patient-panel"><div class="patient-panel-head"><span class="title">Import daily cancellations report</span></div>';
  h+='<input type="file" id="report-file" accept=".csv,.txt,.xlsx,.xls" style="display:none" onchange="importFile(this.files[0])"/>';
  h+='<div style="display:flex;gap:12px;align-items:center;margin-bottom:8px;"><button onclick="document.getElementById(\'report-file\').click()" style="border:none;background:var(--teal);color:#fff;border-radius:6px;padding:9px 18px;font-size:13px;font-weight:600;cursor:pointer;">Choose file</button><span style="font-size:12px;color:var(--muted);" id="file-name">No file chosen</span></div>';
  h+='<div class="paste-hint">Download the cancellations report from Insta HMS (.csv or .xlsx) and choose it here. Patients are sorted into the right location automatically; anyone already showing as Booked is pre-ticked as recouped. Imports go into '+fmtDateHuman(state.date)+'.</div>';
  h+='<details><summary style="font-size:12px;color:var(--teal);cursor:pointer;">Or paste the report text instead</summary><div class="paste-row" style="margin-top:8px;"><textarea id="csv-area" placeholder="Paste the full report here, including its header row" style="min-height:70px;"></textarea><button onclick="importPaste()">Import</button></div></details></div>';

  let tc=0,tr=0,tp=0,tw=0,tf=0,twl=0,rows="";
  LOCATIONS.forEach(l=>{
    const cancs=state.data.cancellations.filter(p=>p.location===l.key);
    const c=state.data.counters[l.key]||{phone:0,whatsapp:0,referral:0,waitlist:0};
    const rate=recoupRate(cancs); const carried=cancs.length; const rn=recoupCount(cancs); const nt=newTotal(c);
    tc+=carried;tr+=rn;tp+=c.phone||0;tw+=c.whatsapp||0;tf+=c.referral||0;twl+=c.waitlist||0;
    rows+="<tr><td>"+l.name+"</td><td>"+carried+"</td><td>"+rn+"</td><td>"+(rate===null?"—":rate+"%")+"</td><td>"+(c.phone||0)+"</td><td>"+(c.whatsapp||0)+"</td><td>"+(c.referral||0)+"</td><td>"+(c.waitlist||0)+"</td><td>"+nt+"</td></tr>";
  });
  const tr2=tc?Math.round(tr/tc*100):null; const gnt=tp+tw+tf+twl;
  rows+='<tr class="total-row"><td>All locations</td><td>'+tc+'</td><td>'+tr+'</td><td>'+(tr2===null?"—":tr2+"%")+'</td><td>'+tp+'</td><td>'+tw+'</td><td>'+tf+'</td><td>'+twl+'</td><td>'+gnt+'</td></tr>';
  h+='<div class="cards-row">'+card("Carried from yesterday",tc,"")+card("Recouped",tr,"",true)+card("Recoup rate",tr2===null?"—":tr2+"%","")+card("New appointments today",gnt,"",true)+'</div>';
  h+='<div class="section-title">By location</div>';
  h+='<table class="summary"><thead><tr><th>Location</th><th>Carried</th><th>Recouped</th><th>Recoup rate</th><th>Phone</th><th>WhatsApp</th><th>Referral</th><th>Waitlist</th><th>New total</th></tr></thead><tbody>'+rows+'</tbody></table>';
  return h;
}

function renderLocation(locKey){
  const loc=LOCATIONS.find(l=>l.key===locKey);
  const cancs=state.data.cancellations.filter(p=>p.location===locKey);
  const c=state.data.counters[locKey]||{phone:0,whatsapp:0,referral:0,waitlist:0};
  const carried=cancs.length; const rn=recoupCount(cancs); const rate=recoupRate(cancs); const nt=newTotal(c);
  let h='<div class="cards-row">'+card("Carried from yesterday",carried,"")+card("Recouped",rn,"",true)+card("Recoup rate",rate===null?"—":rate+"%",rate!==null&&rate>=80?"On target":"")+card("New appointments today",nt,"",true)+'</div>';
  h+='<div class="patient-panel"><div class="patient-panel-head"><span class="title">Cancellations / no-shows — '+loc.name+'</span></div>';
  h+='<div class="paste-row"><textarea id="paste-'+locKey+'" placeholder="Paste patient names from today\'s cancellations report, one per line"></textarea><button onclick="submitPaste(\''+locKey+'\')">Add list</button></div>';
  h+='<div class="paste-hint">Copy the patient name column and paste above, or add one at a time below.</div>';
  h+='<div class="paste-row"><input id="single-'+locKey+'" type="text" placeholder="Patient name" style="flex:1;border:1px solid var(--border);border-radius:6px;padding:8px 10px;font-size:13px;" onkeydown="if(event.key===\'Enter\'){addSingle(\''+locKey+'\');}"/><button class="add-one-btn" onclick="addSingle(\''+locKey+'\')">Add</button></div>';
  h+='<div class="patient-list">';
  if(cancs.length){ cancs.forEach(p=>{
    const meta=[p.mr_no,p.consultant,p.reason].filter(Boolean).join(" · ");
    const digits=(p.phone||"").replace(/[^0-9+]/g,"");
    h+='<div class="patient-row'+(p.recouped?' recouped':'')+'"><input type="checkbox" '+(p.recouped?'checked':'')+' onchange="toggle(\''+p.id+'\')"/>';
    h+='<span style="flex:1;"><div class="p-name">'+esc(p.name)+'</div>'+(meta?'<div style="font-size:11px;color:var(--muted);margin-top:2px;">'+esc(meta)+'</div>':'')+'</span>';
    h+='<input class="p-phone" type="text" placeholder="Phone number" value="'+esc(p.phone||"")+'" onchange="savePhone(\''+p.id+'\',this.value)"/>';
    if(digits) h+='<a class="p-call" href="tel:'+digits+'">&#9742;</a>';
    h+='<input class="p-note" type="text" placeholder="Note" value="'+esc(p.note||"")+'" onchange="saveNote(\''+p.id+'\',this.value)"/>';
    h+='<button class="p-remove" onclick="removePatient(\''+p.id+'\')">&#10005;</button></div>';
  }); } else { h+='<div style="font-size:13px;color:var(--muted);text-align:center;padding:12px 0;">No cancellations or no-shows added for this location yet today.</div>'; }
  h+='</div></div>';
  h+='<div class="section-title">Tap to log</div><div class="counter-grid">';
  COUNTER_DEFS.forEach(cd=>{ h+='<div class="counter-card"><div class="counter-top"><span>'+cd.name+'</span><span class="count">'+(c[cd.field]||0)+'</span></div><div class="counter-buttons"><button class="btn-minus" onclick="adjustCounter(\''+locKey+'\',\''+cd.field+'\','-1)">&#8722;</button><button class="btn-plus" onclick="adjustCounter(\''+locKey+'\',\''+cd.field+'\',1)">+ 1</button></div></div>'; });
  h+='</div>';
  h+='<div class="section-title">Activity log — '+loc.name+'</div><div class="log-box">';
  const logs=(state.data.logs[locKey]||[]);
  if(logs.length){ logs.forEach(e=>{ h+='<div class="log-row"><div class="log-time">'+fmtTime(e.ts)+'</div><div>'+esc(e.label)+'</div></div>'; }); } else { h+='<div style="color:var(--muted);font-size:13px;padding:16px 0;text-align:center;">No activity logged yet today.</div>'; }
  h+='</div>';
  return h;
}

document.getElementById("report-file"); // keep ref pattern
load();
</script>
</body>
</html>"""


@app.route("/")
def index():
    html = (
        INDEX_HTML
        .replace("__LOCATIONS__", json.dumps(LOCATIONS))
        .replace("__COUNTER_DEFS__", json.dumps(COUNTER_DEFS))
    )
    return Response(html, content_type="text/html")


if __name__ == "__main__":
    init_db()
    print("Dashboard running at http://127.0.0.1:5000")
    app.run(debug=True, port=5000)